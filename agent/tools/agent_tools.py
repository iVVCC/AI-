import os
import random
import re
import time
from http.client import responses

import requests
from altair import param

from utils.config_handler import agent_conf
from utils.logger_handler import logger
from utils.path_tool import get_abs_path
from langchain_core.tools import tool
from rag.rag_service import RagSummarizeService
from openpyxl import load_workbook


rag = RagSummarizeService()
user_ids=["1000","1001","1002","1003","1004","1005","1006","1007","1008","1009"]
month_arr = ["2025-01","2025-02","2025-03","2025-04","2025-05","2025-06","2025-07","2025-08","2025-09","2025-10","2025-11","2025-12"]
external_data={}

@tool(description="从向量存储中检索参考资料")
def rag_summarize(query:str)->str:
    return rag.rag_summarize(query)

# @tool(description="获取指定城市的天气，以消息字符串的形式返回")
# def get_weather(city:str)->str:
#     return f"城市{city}天气为晴天，温度为25度"

# @tool(description="获取用户所在城市的名称，以纯字符串形式返回")
# def get_user_location()->str:
#     return random.choice(["偃师","深圳","杭州"])

@tool(description="获取用户的ID，以纯字符串形式返回")
def get_user_id()->str:
    return random.choice(user_ids)

@tool(description="获取当前月份，以纯字符串形式返回")
def get_current_month()-> str:
    return random.choice(month_arr)

def generatr_external_data():
    external_path = get_abs_path(agent_conf["external_records_data"])
    if not os.path.exists(external_path):
        raise FileNotFoundError(f"外部数据{external_path}文件不存在")

    with open(external_path,"r",encoding="utf-8") as f:
        for line in f.readlines()[1:]:
            arr:list[str] = line.strip().split(",")

            user_id: str = arr[0].replace('"',"")
            feature: str = arr[1].replace('"',"")
            efficiency: str = arr[2].replace('"',"")
            consumables: str = arr[3].replace('"',"")
            comparison: str = arr[4].replace('"',"")
            time: str = arr[5].replace('"',"")
            if user_id not in external_data:
                external_data[user_id] = {}
            external_data[user_id][time]={
                "特征":feature,
                "效率":efficiency,
                "消耗":consumables,
                "对比":comparison
            }

@tool(description="从外部系统中获取用户在指定月份的使用记录，以纯字符串形式返回，如果未检索到返回空字符串")
def fetch_external_data(user_id:str,month:str)->str:
    generatr_external_data()
    try:
        return external_data[user_id][month]
    except KeyError:
        logger.warning(f"[fetch_external_data]未能检索到用户:{user_id}在{month}的记录")
        return ""


@tool(description="获取本机公网IP 地址和地理位置信息")
def get_ip_info() -> str:
    """
    获取本机公网IP 和地理位置（使用多个备用服务，提高稳定性）

    Returns:
        str: 包含 IP 地址、地理位置（中文）、运营商等信息的字符串
    """
    # 备用服务列表 - 优先使用支持中文的服务
    ip_services = [
        {
            "name": "ip-api.com",
            "url": "http://ip-api.com/json/",
            "params": "?fields=status,country,countryCode,regionName,city,isp,query,lat,lon&lang=zh-CN",
            "parse": lambda data: (
                f"公网IP: {data.get('query', '未知')}\n"
                f"地理位置：{data.get('country', '未知')} {data.get('regionName', '')} {data.get('city', '')}\n"
                f"运营商：{data.get('isp', '未知')}"
            ) if data.get('status') == 'success' else None
        },
        {
            "name": "ip.sb",
            "ip_url": "https://api.ip.sb/ip",
            "info_url": "https://api.ip.sb/geoip/{ip}",
            "parse": lambda data: (
                f"公网IP: {data.get('ip', '未知')}\n"
                f"地理位置：{data.get('country', '未知')} {data.get('region', '')} {data.get('city', '')}\n"
                f"运营商：{data.get('isp', '未知')}"
            ) if data else None
        },
        {
            "name": "ipapi.com",
            "url": "https://ipapi.co/json/",
            "parse": lambda data: (
                f"公网IP: {data.get('ip', '未知')}\n"
                f"地理位置：{data.get('country_name', '未知')} {data.get('region', '')} {data.get('city', '')}\n"
                f"运营商：{data.get('org', '未知')}"
            ) if data and data.get('error') is None else None
        }
    ]

    # 依次尝试各个服务
    for service in ip_services:
        try:
            # 构建请求 URL
            if service["name"] == "ip-api.com":
                # ip-api.com 需要特殊处理，直接返回完整信息
                full_url = service["url"] + service["params"]
                response = requests.get(full_url, timeout=5)
            elif service["name"] == "ipapi.com":
                # ipapi.com 也直接返回完整信息
                response = requests.get(service["url"], timeout=5)
            else:
                # ip.sb 需要先获取 IP，再获取详细信息
                ip_response = requests.get(service["ip_url"], timeout=5)
                if ip_response.status_code != 200:
                    continue
                ip = ip_response.text.strip()
                info_url = service["info_url"].format(ip=ip)
                response = requests.get(info_url, timeout=5)
            
            if response.status_code != 200:
                continue

            data = response.json()

            # 解析结果
            result = service["parse"](data)
            if result:
                logger.info(f"[get_ip_info] 使用服务 {service['name']} 成功")
                return result

        except requests.RequestException as e:
            logger.warning(f"[get_ip_info] 服务 {service['name']} 失败：{str(e)}")
            continue
        except Exception as e:
            logger.warning(f"失败：{e}")
            time.sleep(2)
            continue

    # 所有服务都失败
    return "获取 IP 信息失败，请检查网络连接"

@tool(description="根据所在区域到城市编码表中查询出城市的 adcode")
def get_city_adcode(city_name:str)->str:
    """
    使用 openpyxl 读取 Excel 获取城市编码
    """
    try:
        file_path = get_abs_path(r"data\AMap_adcode_citycode\AMap_adcode_citycode.xlsx")
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在：{file_path}")
            return ""
        
        # 只读模式打开，避免文件被锁定
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # 从第二行开始遍历（假设第一行是标题）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
                
            city = str(row[0]).strip() if row[0] else ""
            code = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            if city == city_name:
                wb.close()
                return code

        wb.close()
        logger.warning(f"未找到城市 '{city_name}' 的编码")
        return ""

    except Exception as e:
        logger.error(f"读取失败：{e}", exc_info=True)
        return ""

@tool(description="根据获取到的adcode，调用天气情况")
def get_weather_by_adcode(adcode_or_city:str):
    """
    调用高德地图 API 获取天气情况
    参数可以是 adcode 或城市名称（中文）
    """
    try:
        url = f"https://restapi.amap.com/v3/weather/weatherInfo?city={adcode_or_city}"
        parameters={"city": adcode_or_city,
                    "key":"a0e23f009e1256b0ed6912fc0a9311bb"}
        response = requests.get(url, parameters)
        result = response.json()
        
        # # 解析并格式化返回结果
        # if result.get('status') == '1' and result.get('lives'):
        #     live = result['lives'][0]
        #     return f"{live.get('province')}{live.get('city')} - 天气：{live.get('weather')}, 温度：{live.get('temperature')}°C, 风向：{live.get('winddirection')} {live.get('windpower')}级，湿度：{live.get('humidity')}%"
        #
        return result.json()

    except Exception as e:
        logger.error(f"调用失败：{e}")
        return f"获取天气失败：{str(e)}"

@tool(description="无入参，无返回值，调用后触发中间件自动为报告生成的场景动态注入上下文信息，为后续提示词切换提供上下文信息")
def fill_context_for_report():
    return "fill_context_for_report已调用"



if __name__ == '__main__':
    print(fetch_external_data("1010","2025-02"))